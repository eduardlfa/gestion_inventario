from flask import Flask, render_template, request, redirect, url_for, send_file
import sqlite3
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'edu7701')

DATABASE_URL = os.getenv('DATABASE_URL', 'empresa_3d.db')

def get_db_connection():
    conn = sqlite3.connect(DATABASE_URL)
    conn.row_factory = sqlite3.Row
    return conn

def crear_tablas():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS productos (
                    id INTEGER PRIMARY KEY,
                    nombre TEXT NOT NULL,
                    stock INTEGER NOT NULL,
                    precio REAL NOT NULL
                    )''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS compras (
                    id INTEGER PRIMARY KEY,
                    producto_id INTEGER,
                    cantidad INTEGER,
                    fecha TEXT,
                    FOREIGN KEY (producto_id) REFERENCES productos (id)
                    )''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS ventas (
                    id INTEGER PRIMARY KEY,
                    producto_id INTEGER,
                    cantidad INTEGER,
                    fecha TEXT,
                    FOREIGN KEY (producto_id) REFERENCES productos (id)
                    )''')
    conn.commit()
    conn.close()

crear_tablas()

@app.route('/')
def index():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM productos')
    productos = cursor.fetchall()
    conn.close()
    return render_template('index.html', productos=productos)

@app.route('/add_product', methods=['POST'])
def add_product():
    nombre = request.form['nombre']
    stock = int(request.form['stock'])
    precio = float(request.form['precio'])

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, stock FROM productos WHERE nombre = ?', (nombre,))
    producto = cursor.fetchone()

    if producto:
        nuevo_stock = producto[1] + stock
        cursor.execute('UPDATE productos SET stock = ?, precio = ? WHERE id = ?', (nuevo_stock, precio, producto[0]))
    else:
        cursor.execute('INSERT INTO productos (nombre, stock, precio) VALUES (?, ?, ?)', (nombre, stock, precio))
    
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/registrar_compra', methods=['POST'])
def registrar_compra():
    producto_id = int(request.form['producto_id'])
    cantidad = int(request.form['cantidad'])
    fecha = datetime.now().strftime('%Y-%m-%d %H:%M')

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('INSERT INTO compras (producto_id, cantidad, fecha) VALUES (?, ?, ?)', (producto_id, cantidad, fecha))
    cursor.execute('UPDATE productos SET stock = stock + ? WHERE id = ?', (cantidad, producto_id))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

def generar_factura(producto_id, cantidad, fecha):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT nombre, precio FROM productos WHERE id = ?', (producto_id,))
    producto = cursor.fetchone()
    conn.close()
    
    fecha_str = datetime.now().strftime('%Y-%m-%d_%H-%M')
    nombre_factura = f"Factura({fecha_str}).pdf"

    c = canvas.Canvas(ruta_factura, pagesize=A4)
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(300, 800, "HZ Impresiones 3D")
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(300, 780, "FACTURA")

    c.setFont("Helvetica", 12)
    c.drawString(50, 750, f"Fecha: {fecha}")
    c.drawString(50, 730, "Detalles de la venta:")
    c.drawString(50, 710, "-" * 100)
    c.drawString(50, 690, "Producto")
    c.drawString(250, 690, "Cantidad")
    c.drawString(350, 690, "Precio Unitario")
    c.drawString(450, 690, "Total")
    c.drawString(50, 670, "-" * 100)

    y = 650
    subtotal = cantidad * producto['precio']
    c.drawString(50, y, producto['nombre'])
    c.drawString(250, y, str(cantidad))
    c.drawString(350, y, f"Q {producto['precio']:.2f}")
    c.drawString(450, y, f"Q {subtotal:.2f}")

    c.drawString(50, y - 40, "-" * 100)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(350, y - 60, "Total a pagar:")
    c.drawString(450, y - 60, f"Q {subtotal:.2f}")

    c.save()
    return ruta_factura

@app.route('/registrar_venta', methods=['POST'])
def registrar_venta():
    producto_id = int(request.form['producto_id'])
    cantidad = int(request.form['cantidad'])

    print(f"Registrando venta: Producto ID {producto_id}, Cantidad {cantidad}")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT stock FROM productos WHERE id = ?', (producto_id,))
    stock_actual = cursor.fetchone()

    if stock_actual and stock_actual[0] >= cantidad:
        fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute('INSERT INTO ventas (producto_id, cantidad, fecha) VALUES (?, ?, ?)', (producto_id, cantidad, fecha))
        cursor.execute('UPDATE productos SET stock = stock - ? WHERE id = ?', (cantidad, producto_id))
        conn.commit()

        factura_path = generar_factura(producto_id, cantidad, fecha)
        conn.close()

        return redirect(url_for('index', factura_path=factura_path))
    else:
        conn.close()
        print(f"Error: No hay suficiente stock para realizar la venta. Stock disponible: {stock_actual[0] if stock_actual else 'Producto no encontrado'}")
        return redirect(url_for('index'))


@app.route('/download_factura/<path:filename>', methods=['GET'])
def download_factura(filename):
    try:
        ruta_factura = os.path.abspath(filename)
        return send_file(ruta_factura, as_attachment=True)
    except Exception as e:
        return "Error al descargar el archivo", 500



@app.route('/eliminar_producto', methods=['POST'])
def eliminar_producto():
    producto_id = int(request.form['producto_id'])

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM productos WHERE id = ?', (producto_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('index'))

@app.route('/mostrar_inventario')
def mostrar_inventario():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM productos')
    productos = cursor.fetchall()
    conn.close()
    return render_template('inventario.html', productos=productos)

@app.route('/informe')
def informe():
    c = canvas.Canvas("informe_HZ_movimientos.pdf", pagesize=A4)
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(300, 800, "Informe Compras y Ventas - HZ Impresiones 3D")
    c.setFont("Helvetica", 10)
    c.drawCentredString(300, 780, f"Fecha: {datetime.now().strftime('%Y-%m-%d')}")

    y = 750
    c.drawString(100, y, "Compras realizadas:")
    y -= 20
    c.drawString(100, y, "ID|            Producto           | Cantidad | Fecha")
    y -= 20

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''SELECT compras.id, productos.nombre, compras.cantidad, compras.fecha 
                      FROM compras 
                      JOIN productos ON compras.producto_id = productos.id''')
    
    for row in cursor.fetchall():
        c.drawString(100, y, f"{row['id']} | {row['nombre']: <30} | {row['cantidad']} | {row['fecha']}")
        y -= 20

    y -= 20
    c.drawString(100, y, "Ventas realizadas:")
    y -= 20
    c.drawString(100, y, "ID|            Producto           | Cantidad | Fecha")
    y -= 20

    cursor.execute('''SELECT ventas.id, productos.nombre, ventas.cantidad, ventas.fecha 
                      FROM ventas 
                      JOIN productos ON ventas.producto_id = productos.id''')

    for row in cursor.fetchall():
        c.drawString(100, y, f"{row['id']} | {row['nombre']: <30} | {row['cantidad']} | {row['fecha']}")
        y -= 20

    conn.close()
    c.save()
    return send_file("informe_HZ_movimientos.pdf", as_attachment=True)

@app.route('/reporte_excel')
def reporte_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventario"
    
    sheet.append([""])
    sheet.append(["INVENTARIO - HZ IMPRESIONES 3D"])
    sheet.append([""])
    
    sheet.merge_cells('A2:D2')
    title_cell = sheet['A2']
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.font = Font(bold=True, size=14)

    encabezados = ["ID", "Nombre", "Stock", "Precio (Q)"]
    sheet.append(encabezados)

    for col in range(1, len(encabezados) + 1):
        sheet.cell(row=4, column=col).font = Font(bold=True)

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM productos')
    productos = cursor.fetchall()

    for producto in productos:
        sheet.append([producto[0], producto[1], producto[2], producto[3]])

    for row in sheet.iter_rows(min_row=5, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = 'Q #,##0.00'

    for row in sheet.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    workbook.save("reporte_inv.xlsx")
    return send_file("reporte_inv.xlsx", as_attachment=True)

if __name__ == '__main__':
    app.run(debug=False)
